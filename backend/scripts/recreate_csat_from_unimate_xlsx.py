"""
TB_CSAT_SCORE(모의고사) 테이블을 삭제하고 Uni-Mate 데이터 엑셀의 CSAT_SCORE 시트로 재생성합니다.
openpyxl만 사용합니다 (대용량 xlsx에서 pandas 대비 부담이 적음).
"""
from __future__ import annotations

import argparse
import sqlite3
from pathlib import Path

from openpyxl import load_workbook


def _nv(value: object) -> int | float | str | None:
    if value is None:
        return None
    if isinstance(value, str) and value.strip() == "":
        return None
    return value


def load_rows(xlsx_path: Path, sheet_name: str, column_names: list[str]) -> list[tuple]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"시트 없음: {sheet_name!r}. 사용 가능: {wb.sheetnames}")
    ws = wb[sheet_name]
    header_row_no = 1
    header_row = None
    for row_no, candidate in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        names = {str(cell).strip() for cell in candidate if cell is not None}
        if {"csat_id", "student_id"}.issubset(names):
            header_row_no = row_no
            header_row = candidate
            break
    if header_row is None:
        raise SystemExit("CSAT_SCORE 헤더 행을 찾지 못했습니다. csat_id/student_id 컬럼이 필요합니다.")
    header_to_idx = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        name = str(cell).strip()
        if name:
            header_to_idx[name] = idx
    missing = [c for c in column_names if c not in header_to_idx]
    if missing:
        print("missing_optional_columns", missing)

    id_col = header_to_idx.get("csat_id")
    student_col = header_to_idx.get("student_id")
    if id_col is None or student_col is None:
        raise SystemExit("CSAT_SCORE에는 csat_id/student_id 컬럼이 필요합니다.")

    rows: list[tuple] = []
    max_col = max(header_to_idx.values()) + 1
    for r in ws.iter_rows(min_row=header_row_no + 1, max_col=max_col, values_only=True):
        if r is None:
            continue
        has_required_values = id_col < len(r) and student_col < len(r) and r[id_col] is not None and r[student_col] is not None
        if not has_required_values:
            continue
        tup = tuple(_nv(r[header_to_idx[c]]) if c in header_to_idx and header_to_idx[c] < len(r) else None for c in column_names)
        rows.append(tup)
    wb.close()
    return rows


def main() -> int:
    default_xlsx = Path(
        r"E:\sangyun\200 AI기획(20251223~20260514)\프로젝트진행\70 프로세스 설계\Uni-Mate Data_모의고사 20260429.xlsx"
    )
    parser = argparse.ArgumentParser(description="TB_CSAT_SCORE 재생성 (엑셀 CSAT_SCORE)")
    parser.add_argument(
        "--excel",
        type=Path,
        default=default_xlsx,
        help="Uni-Mate 모의고사 엑셀 경로",
    )
    parser.add_argument(
        "--db",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "data" / "uni_mate.db",
        help="SQLite DB 경로",
    )
    parser.add_argument(
        "--sheet",
        default="CSAT_SCORE",
        help="시트 이름",
    )
    args = parser.parse_args()

    xlsx = args.excel
    if not xlsx.is_file():
        raise SystemExit(f"엑셀 파일 없음: {xlsx}")

    schema_columns: list[tuple[str, str]] = [
        ("csat_id", "INTEGER PRIMARY KEY"),
        ("student_id", "INTEGER NOT NULL"),
        ("school_year", "INTEGER"),
        ("exam_year", "INTEGER"),
        ("exam_type", "TEXT"),
        ("exam_month", "INTEGER"),
        ("inquiry_type", "TEXT"),
        ("korean_grade", "INTEGER"),
        ("math_grade", "INTEGER"),
        ("english_grade", "INTEGER"),
        ("korean_history", "INTEGER"),
        ("social_grade", "INTEGER"),
        ("life_and_ethics", "INTEGER"),
        ("ethics_and_thought", "INTEGER"),
        ("korean_geography", "INTEGER"),
        ("world_geography", "INTEGER"),
        ("east_asian_history", "INTEGER"),
        ("world_history", "INTEGER"),
        ("economics", "INTEGER"),
        ("politics_and_law", "INTEGER"),
        ("society_and_culture", "INTEGER"),
        ("science_grade", "INTEGER"),
        ("physics_1", "INTEGER"),
        ("chemistry_1", "INTEGER"),
        ("earth_science_1", "INTEGER"),
        ("life_science_1", "INTEGER"),
        ("physics_2", "INTEGER"),
        ("chemistry_2", "INTEGER"),
        ("earth_science_2", "INTEGER"),
        ("life_science_2", "INTEGER"),
        ("language2_grade", "INTEGER"),
        ("german_1", "INTEGER"),
        ("french_1", "INTEGER"),
        ("spanish_1", "INTEGER"),
        ("chinese_1", "INTEGER"),
        ("japanese_1", "INTEGER"),
        ("russian_1", "INTEGER"),
        ("vietnamese_1", "INTEGER"),
        ("arabic_1", "INTEGER"),
        ("classical_chinese_1", "INTEGER"),
        ("total_score", "REAL"),
        ("percentile", "REAL"),
    ]
    column_names = [name for name, _ in schema_columns]

    print("loading_excel", xlsx)
    rows = load_rows(xlsx, args.sheet, column_names)
    print("excel_data_rows", len(rows))

    insert_sql = (
        "INSERT INTO TB_CSAT_SCORE (" + ", ".join(column_names) + ") VALUES (" + ", ".join(["?"] * len(column_names)) + ")"
    )
    bak_name = "TB_CSAT_SCORE_BAK_20260429_RECREATE"

    with sqlite3.connect(args.db, timeout=60) as conn:
        cur = conn.cursor()
        cur.execute("PRAGMA busy_timeout = 60000")
        cur.execute("PRAGMA foreign_keys = OFF")

        cur.execute(f"DROP TABLE IF EXISTS {bak_name}")
        if cur.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name=? LIMIT 1",
            ("TB_CSAT_SCORE",),
        ).fetchone():
            cur.execute(f"CREATE TABLE {bak_name} AS SELECT * FROM TB_CSAT_SCORE")

        cur.execute("DROP TABLE IF EXISTS TB_CSAT_SCORE")
        create_sql = (
            "CREATE TABLE TB_CSAT_SCORE ("
            + ", ".join([f"{name} {dtype}" for name, dtype in schema_columns])
            + ", FOREIGN KEY (student_id) REFERENCES TB_STUDENT_PROFILE(student_id)"
            + ")"
        )
        cur.execute(create_sql)
        cur.executemany(insert_sql, rows)
        conn.commit()

        cur.execute("PRAGMA foreign_keys = ON")

        table_count = cur.execute("SELECT COUNT(1) FROM TB_CSAT_SCORE").fetchone()[0]
        sample = cur.execute(
            """
            SELECT csat_id, student_id, school_year, exam_year, exam_type, exam_month, inquiry_type, total_score, percentile
            FROM TB_CSAT_SCORE
            ORDER BY csat_id
            LIMIT 5
            """
        ).fetchall()
        cols = [r[1] for r in cur.execute("PRAGMA table_info(TB_CSAT_SCORE)").fetchall()]

    print("inserted_rows", len(rows))
    print("table_count", table_count)
    print("column_count", len(cols))
    print("sample", sample)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
